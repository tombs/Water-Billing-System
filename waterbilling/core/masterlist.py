# This module uses OpenPyxl.
# OpenPyxl is a Python library to read/write Excel 2007 xlsx/xlsm files.


# To Install:
#
# pip install openpyxl


# Import OpenPyxl modules
from openpyxl.reader.excel import load_workbook
from openpyxl.style import Border
from openpyxl import Workbook
from openpyxl.style import Color, Fill
from openpyxl.cell import Cell
from django.db.models import Sum, Max, F

# Import other modules
from django.core.files.storage import FileSystemStorage


#Import Config from models
from .models import Config, Bill, BillingSchedule, FinancialTransaction, Account


# Define the MasterList class
class WbsMasterList:
    '''

    This class creates the Water Billing System Masterlist. 
    This creates an output file which is in an Excel (xlsx) format.
    It aims to ease the creation / recreation of the MasterList at any given time

    '''
  
    filestorage = None
    filepath = ""
    filename = ""
    filepath_and_name = ""
    workbook = None
    currentrow = 0
    currentcol = 0
    maxcol = 31
    current_bill_period = None
    previous_bill_period = None
    period_covered = ""
    due_date = ""
    phase = ""
    data = None

    def __init__(self, filename="masterlist.xlsx",filepath="", bill_period=None):
        '''

        Initialize the MasterList class stuff here

        '''

        if filepath!="":
            self.filepath = filepath
        else:
            self.filestorage = FileSystemStorage()
            self.filepath = self.filestorage.location
        
        self.filename = filename
        self.filepath_and_name = self.filepath + "/" + self.filename

        
        # Open the Masterlist workbook from the default Django storage location
        # If the Masterlist workbook is missing, this will create a new MasterList workbook
        try:
            self.open()   
        except Exception, e:
            print "Error opening master list: " + str(e)
            print "Creating a new WBS master list... "
            self.new()


        # Load the default configurations needed.
        # This is assumed to be existing in the Config Table.
        self.load_configs()


        #Get the current period based on the Business Date in the Config table.
        #Alternatively, a BillingSchedule object can be passed as 'bill_period' here
        if bill_period == None:
            from datetime import datetime
            today = datetime.strptime(self.business_date,"%Y-%m-%d") 
            self.current_bill_period =   BillingSchedule.objects.get(start_date__lte=today, end_date__gte=today)   
        else:
            self.current_bill_period = bill_period

        # Get the previous bill period based on the current bill period
        self.previous_bill_period = self.current_bill_period.get_previous_by_start_date()

        self.period_covered = str(self.current_bill_period)
        self.due_date = self.current_bill_period.due_date()

        
    
    def new(self):
        
        ''' 
        
        Create a new workbook
        
        '''
        
        self.workbook = Workbook()

    
    def open(self):
        
        '''

        Open a workbook using the class filename

        '''

        self.workbook = load_workbook(self.filepath_and_name)

    
    def save(self):
        '''

        save workbook using class filename

        '''

        self.workbook.save(self.filepath_and_name)

    def active_sheet(self):
        '''

        return the currently active sheet in the workbook

        '''

        return self.workbook.get_active_sheet()

    def set_border(self, cell_range, ws=None):
        
        '''
        
        creates a border for a range of cells
        
        ws         :    Workbook sheet
        cell_range :    "A1:H1000"   
                
        '''
        
        if ws==None:
            ws = self.workbook.get_active_sheet()

        rows = ws.range(cell_range)
        for row in rows:
            row[0].style.borders.left.border_style = Border.BORDER_THIN
            row[-1].style.borders.right.border_style = Border.BORDER_THIN
        for c in rows[0]:
            c.style.borders.top.border_style = Border.BORDER_THIN
        for c in rows[-1]:
            c.style.borders.bottom.border_style = Border.BORDER_THIN


    def set_border_by_cell(self, row, column, ws):
        
        '''
        
        creates a border for a single cell
        
        ws         :    Workbook sheet
        cell_range :    "A1:H1000"   
                
        '''
        
            

        d = ws.cell(row = row, column = column)

        d.style.borders.left.border_style = Border.BORDER_THIN
        d.style.borders.right.border_style = Border.BORDER_THIN
        d.style.borders.top.border_style = Border.BORDER_THIN
        d.style.borders.bottom.border_style = Border.BORDER_THIN


    def clear_border(self, cell_range, ws=None):
        
        '''
        
        removes the borders from a range of cells
        
        ws         :    Workbook sheet
        cell_range :    "A1:H1000"   
                
        '''
        
        if ws==None:
            ws = self.workbook.get_active_sheet()

        rows = ws.range(cell_range)
        for row in rows:
            for column in row:
                column.style.borders.left.border_style = Border.BORDER_NONE            
                column.style.borders.right.border_style = Border.BORDER_NONE
                column.style.borders.top.border_style = Border.BORDER_NONE
                column.style.borders.bottom.border_style = Border.BORDER_NONE


    def paint_cells(self, cell_range, color="BLUE", ws=None):
        
        '''
        
        this paints a range of cells using a specified color

        ws         :    Workbook sheet
        cell_range :    "A1:H1000"   
        color      :    Color to paint the cells

        NOTE:  There is currently an issue with the WHITE color, so defaulting to "BLUE"
                
        '''

            
        if ws==None:
            ws = self.workbook.get_active_sheet()


        rows = ws.range(cell_range)
        for row in rows:
            for column in row:
                column.style.fill.fill_type = Fill.FILL_SOLID
                column.style.fill.start_color.index = getattr(Color,color)


    def load_configs(self):
        '''

        This funciton loads the configurations needed for the rendering of the MasterList

        '''

        self.title = Config.objects.get(name='ml_title').value
        self.project = Config.objects.get(name='ml_project').value
        self.location =  Config.objects.get(name='ml_location').value
        self.business_date = Config.objects.get(name='business_date').value
        

    def render(self, bill_period=None, ):
        '''

        This renders the entire MasterList.


        bill_period:   This is the BillingSchedule object to render. 
                       Defaults to the BillingSchedule of the current Business Date.

        '''


        current_bill_period = None

    	if bill_period == None:
            from datetime import date
            today = date.today()
            current_bill_period =   BillingSchedule.objects.get(start_date__lte=today, end_date__gte=today)   
        else:
            current_bill_period = bill_period

        previous_bill_period = current_bill_period.get_previous_by_start_date()

        wbx = self.workbook
    	w = wbx.get_active_sheet()
    	
        
        # Render the Title Area

    	self.paint_cells("A1:AE5")  # paint this section white
    	w["A1"] = self.title
    	w["A2"] = self.project
    	w["A3"] = self.location
        w["A4"] = self.period_covered
        w["A5"] = self.due_date


        self.currentrow = 5


        # Render the data per Phase
        # The render_phase() function is used

        self.phase = "PHASE 1"
        data = self.get_phase_data(self.phase)
        self.data = data
        self.render_phase(data)

        self.phase = "PHASE 2"
        data = self.get_phase_data(self.phase)
        self.data = data
        self.render_phase(data)

        self.phase = "PHASE 3"
        data = self.get_phase_data(self.phase)
        self.data = data
        self.render_phase(data)

        self.phase = "PHASE 4"
        data = self.get_phase_data(self.phase)
        self.data = data
        self.render_phase(data)

        self.phase = "PHASE 5"
        data = self.get_phase_data(self.phase)
        self.data = data
        self.render_phase(data)

        self.phase = "PARKLANE"
        data = self.get_phase_data(self.phase)
        self.data = data
        self.render_phase(data)



    def get_phase_data(self, phase):
        '''

        This function executes a query to retrieve the  Phase Data to be rendered.

        phase:    This is the string value of the Phase to be Rendered. 
                  Must be a value in the Address4 field of the Address Table

        '''

        queryset = FinancialTransaction.objects.filter(
            id__in=Account.objects.annotate(last_transaction_id=Max(
                'financialtransaction')).values_list(
                'last_transaction_id'), 
                balance__gt=0).values(
            'account__id',
            'account__customer__last_name',
            'account__customer__first_name',
            'account__address__address1', 
            'account__address__address4',
            'balance',
            'account__status',
            'account__remarks',
            'account__accountmeter__meter__meter_uid',
            'account__accountmeter__meter__meterread__current_reading',
            'account__accountmeter__meter__meterread__previous_reading',
            'account__accountmeter__meter__meterread__usage',
            'account__bill__current_charge',
            'account__bill__previous_balance',
            'account__bill__amount_due',
            'account__bill__penalty_amount',
              ).order_by('account__address__address4').order_by('account__address__address1')

        return queryset.filter(account__address__address4=phase)


    def parse_address_units(self, address):
        '''

        Parses block and lot of address. 


        '''

        a = address.split(" ")
        return (a[1], a[3])

    def render_phase(self, data):
        '''

        This renders the data on a per phase basis on the masterlist.

        data :  list of accounts per phase

        '''

        row = self.currentrow
        ws = self.workbook.get_active_sheet()

        #Create a boxed set of cells

        #Box 3 rows first for the title rows
        for i in range(0,3):
            for col in range(0,self.maxcol):
                self.set_border_by_cell(row, col, ws)
            row = row + 1

        # Then box the rest based on the number of data
        for record in data:
            for col in range(0,self.maxcol):
                self.set_border_by_cell(row, col, ws)
            row = row + 1


        #reset row to original position
        row = self.currentrow

        # 1st Row of Phase Display
        ws.merge_cells(start_row=row, start_column=0, end_row=row, end_column=3)
        phase_cell = ws.cell(row=row, column=0)
        phase_cell.value = self.phase

        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=22)
        period_cell = ws.cell(row=row, column=4)
        period_cell.value = str(self.current_bill_period)

        ws.merge_cells(start_row=row, start_column=23, end_row=row, end_column=26)
        period_cell = ws.cell(row=row, column=23)
        period_cell.value = str(self.previous_bill_period)

        ws.merge_cells(start_row=row, start_column=27, end_row=row, end_column=29)
        period_cell = ws.cell(row=row, column=27)
        period_cell.value = ""        


        # 2nd Row of Phase Display
        row = row + 1

        cell = ws.cell(row = row, column = 0)
        cell.value = "NO."

        cell = ws.cell(row = row, column = 1)
        cell.value = "BLK"

        cell = ws.cell(row = row, column = 2)
        cell.value = "LOT"

        cell = ws.cell(row = row, column = 3)
        cell.value = "OWNER"

        cell = ws.cell(row = row, column = 4)
        cell.value = "METER"


        ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
        cell = ws.cell(row=row, column=5)
        cell.value = "READING (m3)"


        ws.merge_cells(start_row=row, start_column=7, end_row=row, end_column=8)
        cell = ws.cell(row=row, column=7)
        cell.value = "CONSUMPTION"

        cell = ws.cell(row = row, column = 9)
        cell.value = "RATE  "        

        cell = ws.cell(row = row, column = 10)
        cell.value = "NXT 10  "

        cell = ws.cell(row = row, column = 11)
        cell.value = "NXT 10  "        

        cell = ws.cell(row = row, column = 12)
        cell.value = "NXT 10  "

        cell = ws.cell(row = row, column = 13)
        cell.value = "NXT 10 "

        cell = ws.cell(row = row, column = 14)
        cell.value = "NXT 10  "

        cell = ws.cell(row = row, column = 15)
        cell.value = "NXT 10  "

        cell = ws.cell(row = row, column = 16)
        cell.value = "NXT 10  "

        cell = ws.cell(row = row, column = 17)
        cell.value = "NXT 10  "

        cell = ws.cell(row = row, column = 18)
        cell.value = "NXT 10  "

        cell = ws.cell(row = row, column = 19)
        cell.value = "NXT 10 "

        cell = ws.cell(row = row, column = 20)
        cell.value = "NXT 10  "

        cell = ws.cell(row = row, column = 21)
        cell.value = "NXT 10  "

        cell = ws.cell(row = row, column = 22)
        cell.value = "TOTAL  "

        ws.merge_cells(start_row=row, start_column=23, end_row=row, end_column=25)
        cell = ws.cell(row=row, column=23)
        cell.value = "(COLLECTIBLES)"


        cell = ws.cell(row = row, column = 26)
        cell.value = "TOTAL  "        

        cell = ws.cell(row = row, column = 27)
        cell.value = "AMOUNT  "

        cell = ws.cell(row = row, column = 28)
        cell.value = "DATE  "


        cell = ws.cell(row = row, column = 29)
        cell.value = "OR #  "


        # 3rd Row of Phase Display
        row = row + 1

        cell = ws.cell(row = row, column = 0)
        cell.value = " "

        cell = ws.cell(row = row, column = 1)
        cell.value = " "

        cell = ws.cell(row = row, column = 2)
        cell.value = " "

        cell = ws.cell(row = row, column = 3)
        cell.value = " "

        cell = ws.cell(row = row, column = 4)
        cell.value = "NO. "


        cell = ws.cell(row = row, column = 5)
        cell.value = "PREV. "

        cell = ws.cell(row = row, column = 6)
        cell.value = "PRES. "

        cell = ws.cell(row = row, column = 7)
        cell.value = "MIN. "

        cell = ws.cell(row = row, column = 8)
        cell.value = "EXCESS. "



        cell = ws.cell(row = row, column = 9)
        cell.value = "MIN.  "        

        cell = ws.cell(row = row, column = 10)
        cell.value = "at P16.40  "

        cell = ws.cell(row = row, column = 11)
        cell.value = "at P17.20  "        

        cell = ws.cell(row = row, column = 12)
        cell.value = "at P16 "

        cell = ws.cell(row = row, column = 13)
        cell.value = "at P18 "

        cell = ws.cell(row = row, column = 14)
        cell.value = "at P20  "

        cell = ws.cell(row = row, column = 15)
        cell.value = "at P22  "

        cell = ws.cell(row = row, column = 16)
        cell.value = "at P24  "

        cell = ws.cell(row = row, column = 17)
        cell.value = "at P26  "

        cell = ws.cell(row = row, column = 18)
        cell.value = "at P28   "

        cell = ws.cell(row = row, column = 19)
        cell.value = "at P30  "

        cell = ws.cell(row = row, column = 20)
        cell.value = "at P32 "

        cell = ws.cell(row = row, column = 21)
        cell.value = "at P34  "

        cell = ws.cell(row = row, column = 22)
        cell.value = "AMOUNT  "

        cell = ws.cell(row = row, column = 23)
        cell.value = "Unpaid bill  "

        cell = ws.cell(row = row, column = 24)
        cell.value = "PENALTY "

        cell = ws.cell(row = row, column = 25)
        cell.value = "SUBTOTAL  "


        cell = ws.cell(row = row, column = 26)
        cell.value = "AMOUNT  "        

        cell = ws.cell(row = row, column = 27)
        cell.value = "PAID  "

        cell = ws.cell(row = row, column = 28)
        cell.value = " "


        cell = ws.cell(row = row, column = 29)
        cell.value = " "


        # 4th row onwards -- Display of Data

        row = row + 1

         #Print the data 
        counter = 0
        for record in data:
            #for col in range(0,self.maxcol):
            #    self.set_border_by_cell(row, col, ws)
            
            counter = counter + 1

            blk, lot = self.parse_address_units(record['account__address__address1'])

            cell = ws.cell(row = row, column = 0)
            cell.value = counter

            cell = ws.cell(row = row, column = 1)
            cell.value = blk

            cell = ws.cell(row = row, column = 2)
            cell.value = lot

            cell = ws.cell(row = row, column = 3)
            cell.value = record['account__customer__last_name'] + "," + record['account__customer__first_name']
            print " Customer: %s " % (record['account__customer__last_name'] + "," + record['account__customer__first_name'])
            cell = ws.cell(row = row, column = 4)
            cell.value = record['account__accountmeter__meter__meter_uid']


            cell = ws.cell(row = row, column = 5)
            cell.value = record['account__accountmeter__meter__meterread__previous_reading']


            cell = ws.cell(row = row, column = 6)
            cell.value = record['account__accountmeter__meter__meterread__current_reading']
           

            cell = ws.cell(row = row, column = 7)
            cell.value = record['account__accountmeter__meter__meterread__usage']

            cell = ws.cell(row = row, column = 8)
            cell.value = record['account__accountmeter__meter__meterread__usage'] - 10



            cell = ws.cell(row = row, column = 9)
            cell.value = "156 " # Should be based on the minimum rate, not hard coded        

            cell = ws.cell(row = row, column = 10)
            cell.value = "  "

            cell = ws.cell(row = row, column = 11)
            cell.value = "  "        

            cell = ws.cell(row = row, column = 12)
            cell.value = " "

            cell = ws.cell(row = row, column = 13)
            cell.value = " "

            cell = ws.cell(row = row, column = 14)
            cell.value = "  "

            cell = ws.cell(row = row, column = 15)
            cell.value = "  "

            cell = ws.cell(row = row, column = 16)
            cell.value = "  "

            cell = ws.cell(row = row, column = 17)
            cell.value = "  "

            cell = ws.cell(row = row, column = 18)
            cell.value = "   "

            cell = ws.cell(row = row, column = 19)
            cell.value = "  "

            cell = ws.cell(row = row, column = 20)
            cell.value = " "

            cell = ws.cell(row = row, column = 21)
            cell.value = "  "

            cell = ws.cell(row = row, column = 22)
            cell.value = record['account__bill__current_charge']

            cell = ws.cell(row = row, column = 23)
            cell.value = record['account__bill__previous_balance']

            cell = ws.cell(row = row, column = 24)
            cell.value = record['account__bill__penalty_amount']

            cell = ws.cell(row = row, column = 25)
            try:
                cell.value = record['account__bill__previous_balance'] + record['account__bill__penalty_amount']
            except:
                cell.value = "0.00"
                
            cell = ws.cell(row = row, column = 26)
            try:
                cell.value = record['account__bill__amount_due']        
            except:
                cell.value = "0.00"

            cell = ws.cell(row = row, column = 27)
            cell.value = " "

            cell = ws.cell(row = row, column = 28)
            cell.value = " "


            cell = ws.cell(row = row, column = 29)
            cell.value = " "


            
            row = row + 1
            

        # Set the latest row 2 lines below the last record. 

        row = row + 1
        row = row + 1

        self.currentrow = row

            

                

