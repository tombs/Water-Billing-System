# -*- coding: utf-8 -*-
"""
dataloader.py
Facility to load data into the database

"""

import csv
import sys
from datetime import date, time, datetime
import codecs

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils import timezone


from optparse import make_option
#import argparse

from core import models
from core.utils import get_business_date, business_date_to_date, load_wbs_configs, load_wbs_users
from openpyxl import load_workbook

class BaseLoader(object):

    user = 'batch'

    def __init__(self, path, format):
        pass

    def load(self):
        pass



class CustomerLoader(BaseLoader):
    Model = models.Customer

    def __init__(self, path, format='csv'):
        self.path = path
        self.format = format

    def load(self):

        with open(self.path, 'r') as f:
            if self.format == 'csv': 
                self.reader= csv.DictReader(f)
            else:
                self.reader = []
            data = []
            for row in self.reader:
                data.append(self.Model(**row))
            print "loading", len(data), "for", self.Model
            self.Model.objects.bulk_create(data)


class AddressLoader(BaseLoader):
    Model = models.Address

    def __init__(self, path, format='csv'):
        self.path = path
        self.format = format

    def load(self):

        with open(self.path, 'r') as f:
            if self.format == 'csv': 
                self.reader= csv.DictReader(f)
            else:
                self.reader = []
            data = []
            for row in self.reader:
                row['created_by'] = self.user
                row['last_updated_by'] = self.user
                data.append(self.Model(**row))
            print "loading", len(data), "for", self.Model
            self.Model.objects.bulk_create(data)


class AccountTypeLoader(BaseLoader):
    Model = models.AccountType

    def __init__(self, path, format='csv'):
        self.path = path
        self.format = format

    def load(self):

        with open(self.path, 'r') as f:
            if self.format == 'csv': 
                self.reader= csv.DictReader(f)
            else:
                self.reader = []
            data = []
            for row in self.reader:
                row['created_by'] = self.user
                row['last_updated_by'] = self.user              
                data.append(self.Model(**row))
            print "loading", len(data), "for", self.Model
            self.Model.objects.bulk_create(data)

class RateLoader(BaseLoader):
    Model = models.Rate

    def __init__(self, path, format='csv'):
        self.path = path
        self.format = format

    def load(self):

        with open(self.path, 'r') as f:
            if self.format == 'csv': 
                self.reader= csv.DictReader(f)
            else:
                self.reader = []
            data = []
            for row in self.reader:
                row['created_by'] = self.user
                row['last_updated_by'] = self.user
                data.append(self.Model(**row))
            print "loading", len(data), "for", self.Model
            self.Model.objects.bulk_create(data)

class RateChargeLoader(BaseLoader):
    Model = models.RateCharge

    def __init__(self, path, format='csv'):
        self.path = path
        self.format = format

    def load(self):

        with open(self.path, 'r') as f:
            if self.format == 'csv': 
                self.reader= csv.DictReader(f)
            else:
                self.reader = []
            data = []
            for row in self.reader:
                row['created_by'] = self.user
                row['last_updated_by'] = self.user
                data.append(self.Model(**row))
            print "loading", len(data), "for", self.Model
            self.Model.objects.bulk_create(data)

class BillingScheduleLoader(BaseLoader):
    Model = models.BillingSchedule

    def __init__(self, path, format='csv'):
        self.path = path
        self.format = format

    def load(self):
        load_wbs_configs()
        from datetime import timedelta, datetime
        from dateutil.relativedelta import relativedelta

        Config = models.Config
        read_bill_interval = int(Config.objects.get(name='read_bill_interval').value)
        billing_duration = int(Config.objects.get(name='billing_duration').value)

        with open(self.path, 'r') as f:
            if self.format == 'csv': 
                self.reader= csv.DictReader(f)
            else:
                self.reader = []
            data = []
            for row in self.reader:     
                row['created_by'] = self.user
                row['last_updated_by'] = self.user
                reading_start_date = datetime.strptime(row['reading_start_date'],'%Y-%m-%d').date()
                reading_end_date = datetime.strptime(row['reading_end_date'],'%Y-%m-%d').date()
                start_date = reading_end_date + timedelta(read_bill_interval)
                row['start_date'] = start_date.strftime('%Y-%m-%d')
                end_date =  start_date + relativedelta(months=+billing_duration) - timedelta(days=1)
                row['end_date'] = end_date.strftime('%Y-%m-%d')
                data.append(self.Model(**row))
            print "loading", len(data), "for", self.Model
            self.Model.objects.bulk_create(data)

class MeterLoader(BaseLoader):
    Model = models.Meter

    def __init__(self, path, format='csv'):
        self.path = path
        self.format = format

    def load(self):

        with open(self.path, 'r') as f:
            if self.format == 'csv': 
                self.reader= csv.DictReader(f)
            else:
                self.reader = []

            data = []
            
            for row in self.reader:
                row['created_by'] = self.user
                row['last_updated_by'] = self.user
                data.append(self.Model(**row))
            
            print "loading", len(data), "for", self.Model
            self.Model.objects.bulk_create(data)
   
class AccountLoader(BaseLoader):
    Model = models.Account

    def __init__(self, path, format='csv'):
        self.path = path
        self.format = format

    def load(self):

        with open(self.path, 'r') as f:
            if self.format == 'csv': 
                self.reader= csv.DictReader(f)
            else:
                self.reader = []
            data = []
            for row in self.reader:
                row['created_by'] = self.user
                row['last_updated_by'] = self.user
                data.append(self.Model(**row))
            print "loading", len(data), "for", self.Model
            self.Model.objects.bulk_create(data)

class MeterReadLoader(BaseLoader):
    Model = models.MeterRead

    def __init__(self, path, format='xlsx', headers=True):
        self.path = path
        self.format = format
        self.headers = headers

    def load(self):

        from copy import copy
        load_details = {}
        processed = 0
        ignored = []
        error = []
        skipped = 0
        total = 0
        rows = 0
        if self.format=='csv':

            with open(self.path, 'r') as f:
                if self.format == 'csv': 
                    self.reader= csv.DictReader(f)
                else:
                    self.reader = []
                
                data = []
                errors = []
                period = models.Config.objects.get(name='active_period').value
                bs = models.BillingSchedule.objects.get(pk=period)

                #load_details['rows'] = len(self.reader)

                
                for row in self.reader:
                    rows+=1
                #    row['created_by'] = self.user
                #    row['last_updated_by'] = self.user
                #    data.append(self.Model(**row))
                #print "loading", len(data), "for", self.Model
                #self.Model.objects.bulk_create(data)
                    try:
                        error.append(copy(row))

                        meter_uid = row.pop('meter_uid')
                        meter = models.Meter.objects.get(meter_uid=meter_uid)
                        previous_reading = row.pop('previous_reading')
                        address = row.pop('address').strip()
                        customer = row.pop('customer')
                        last_name, first_name = customer.split(",")
                        last_name = last_name.strip()
                        first_name = first_name.strip()
                        print "last_name: ", last_name
                        print "first_name: ", first_name
                        customer = models.Customer.objects.get(last_name=last_name, first_name=first_name)
                        address = models.Address.objects.get(address1=address)
                        account = models.Account.objects.get(customer=customer, address=address)
                        
                        read_time = row.pop('read_time')
                        if (read_time is None) or (read_time.strip()==''):
                            row['read_time'] = '12:00'
                        #account = 
                        row['meter'] = meter
                        row['account'] = account
                        row['billing_schedule'] = bs
                        print "customer: ", customer
                        print "row: ", row
                        data.append(self.Model(**row))
                        data[-1].save()      
                        processed+=1
                        error.pop()

                    except Exception, e:
                        err = error.pop()
                        print "Skipping row with Error", meter_uid, customer, address, row, e
                        error.append("row: %s  error: % s" % (str(err), str(e)))
                        print "error: ",e
                        continue

            load_details['rows'] = rows
            load_details['processed'] = processed
            load_details['error'] = len(error)
            load_details['skipped'] = skipped
            load_details['ignored'] = len(ignored)
            load_details['total'] = processed + len(ignored) + len(error)

            print "Load Results: " + str(load_details)
            if error:
                print "Errors!"
                print
                print

                for err in error:
                    print err


        elif self.format == 'xlsx':
            print "loading.."
            wb2 = load_workbook(self.path)
            print "loaded!"
            print "sheets: ", wb2.get_sheet_names()
            ws = wb2.get_active_sheet()
            rows = ws.rows
            print "rows: ", rows
            load_details['rows'] = len(rows)

            fields = ['meter_uid', 'customer', 'address', 'read_date', 'read_time','previous_reading','current_reading']

            start = 0
            if self.headers:
                start = 1

            load_details['header'] = start

            period = models.Config.objects.get(name='active_period').value
            bs = models.BillingSchedule.objects.get(pk=period)

            
            for i in range(start, len(ws.rows)):
                print "cell %d:  %s", i, ws.cell(row=i, column=1)
                if (ws.cell(row=i, column=1).value is None) or (ws.cell(row=i, column=1).value.strip() == ''):
                    skipped+=1
                    continue                    
                else:
                    try:
                        row = {}
                        for field in fields:
                            print "field: ", field
                            print "cell: ", ws.cell(row=i,column=fields.index(field)).value
                            #row[i].decode('iso-8859-1').strip()
                            #print "decode: ", ws.cell(row=i,column=fields.index(field)).value.decode
                            row[field] = ws.cell(row=i,column=fields.index(field)).value

                        error.append(copy(row))
                        print "parsing.."
                        
                        if (row['current_reading'] is None):
                            print "Blank current reading.. ignoring row %d , row: %s " % (i, str(row))
                            ignored.append(row)
                            error.pop()
                            continue

                        meter_uid = row.pop('meter_uid')
                        print "meter_uid: ", meter_uid
                        #meter = models.Meter.objects.get(meter_uid=meter_uid)
                        address = row.pop('address').strip()
                        customer = row.pop('customer')
                        last_name, first_name = customer.split(",")
                        last_name = last_name.strip().upper()
                        first_name = first_name.strip().upper()
                        print "last_name: ", last_name
                        print "first_name: ", first_name
                        customer = models.Customer.objects.get(last_name=last_name, first_name=first_name)
                        address = models.Address.objects.get(address1=address)
                        account = models.Account.objects.get(customer=customer, address=address)
                        print "row, before adding.. : ", row
                        
                        account.add_meter_read(current_reading=row['current_reading'],period=bs.id)

                        processed+=1
                        error.pop()
                        #d['account'] = account
                        
                        #data = Model(**d)
                        #data.save()
                    except Exception, e:
                        err = error.pop()
                        error.append("row: %s  error: % s" % (str(err), str(e)))
                        print "error: ",e

            
                        #print "Skipping row with Error", meter_uid, customer, address, e
            load_details['processed'] = processed
            load_details['error'] = len(error)
            load_details['skipped'] = skipped
            load_details['ignored'] = len(ignored)
            load_details['total'] = processed + len(ignored) + len(error)

            print "Load Results: " + str(load_details)
            if error:
                print "Errors!"
                print
                print

                for err in error:
                    print err

class FelizanaLoader(BaseLoader):
    def __init__(self, path, format='csv'):
        self.path = path
        self.format = format

    @transaction.commit_manually
    def load(self,business_date=None):
        addresses = []
        customers = []
        meters = []
        try:
            print "getting schedule and accounttype"
            bs, created = models.BillingSchedule.objects.get_or_create(start_date=date(2014, 2, 16), end_date=date(2014, 3, 15))
            if business_date is None:
                business_date = models.Config.objects.get(name='business_date').value
            #account_type = models.AccountType.objects.get(description = 'Residential')
            print "loading default configs.. "
            load_wbs_configs()
            load_wbs_users()
            print ".. done loading!"
        except Exception, e:
            print e
            transaction.rollback()
            raise e

        with open(self.path, 'r') as f:
            if self.format == 'csv': 
                self.reader= csv.DictReader(f)
            else:
                self.reader = []
            
            #PHASE,BLK,LOT,OWNER,date_installed,meter_uid,previous_reading,current_reading,usage,read_date
            duplicate_meters = []
            for row in self.reader:
                try:
                    print "Loading", row
                    for i in row.keys():
                        row[i] = row[i].decode('iso-8859-1').strip()#.encode('utf-8')
                    print "row", row
                    if not (row['PHASE'] and row['OWNER'] and row['meter_uid']):
                        print "Skipped", row
                        continue

                    # for Account Type
                    account_type = models.AccountType.objects.get(description = row['account_type'])

                    #for Customer
                    if ',' in row['OWNER']:
                        row['last_name'], row['first_name'] = row['OWNER'].split(',',1)
                        row['last_name'] = row['last_name'].strip()
                        row['first_name'] = row['first_name'].strip()

                    else:
                        row['last_name'] = row['OWNER']
                        row['first_name'] = ''
                    customer, created = models.Customer.objects.get_or_create(first_name = row['first_name'].upper(), last_name = row['last_name'].upper() )
                    print "Customer created", created, customer
                    
                    #for Address
                    address1 = u' '.join( [u'Blk', row['BLK'], u'Lot', row['LOT'], row['PHASE'] + u', ' + u'FelizAna Estate Subd., Brgy. Pasong Buaya, Imus, Cavite']).upper()
                    address2 = row['BLK']
                    address3 = row['LOT']
                    address4 = row['PHASE'].upper()
                    phase = ''
                    if ' ' in row['PHASE']:
                        phase = row['PHASE'].split(' ')[1]
                    else:
                        phase = row['PHASE'][0]

                    address, created = models.Address.objects.get_or_create( 
                            address1 = address1, 
                            defaults = {'address4':address4, 
                                        'address2':address2,
                                        'address3':address3,
                                        'zip_code':4103
                                       })
                    print "Address created", created, address
                    #for Meter
                    if not row['date_installed']:
                        row['date_installed'] = business_date_to_date()
                    try:
                        print "old_meter: ", row['meter_uid']
                        
                        meter = models.Meter.objects.filter(meter_uid=row['meter_uid'])
                        print "the meter: ", meter
                        if not meter:
                            meter = models.Meter.objects.create(meter_uid = row['meter_uid'], 
                                date_installed = row['date_installed'], status='active') #date.today())                            
                        else:
                            if row['meter_uid'] not in duplicate_meters:
                                print "adding duplicate meter: ", row['meter_uid']
                                duplicate_meters.append(row['meter_uid'])
                            print "phase: ", phase
                            new_meter = u'-'.join([row['meter_uid'],phase,row['BLK'],row['LOT']])
                            print " and new meter, again: ", new_meter
                            meter = models.Meter.objects.create(meter_uid = new_meter, 
                                date_installed = row['date_installed'], status='active') #date.today())  
                        print "WARN: Must be unique Meter created", meter
                    except Exception, f:
                        print "Cant do this anymore! -->  ", f
                        transaction.rollback()
                        exit()


                    #for Account

                    amount = row['balance']
                    account_status = row['status']
                    
                    account, created = models.Account.objects.get_or_create(customer=customer,
                         address = address, account_type=account_type, 
                         defaults={'status':account_status })
                    print "account created", created, account
                    #MeterRead


                    account.add_account_meter(meter.pk)



                    meterread, created = models.MeterRead.objects.get_or_create(meter = meter, account = account, billing_schedule = bs,
                                    read_date=row['read_date'],
                                    defaults={
                                                'read_time':time(),
                                                'current_reading':row['current_reading'],
                                                'previous_reading':row['previous_reading'],
                                                'usage':row['usage']
                                            })
                    print "meterread", meterread, created

                    account.add_adjustment(amount=amount,adjustment_type='debit',adjustment_date=business_date,description='initial balance')
                except Exception, e:
                    print "Rollback"
                    transaction.rollback()
                    print e
                    raise e
                else:
                    transaction.commit()

            print "duplicate meters!: ", duplicate_meters
            for uid in duplicate_meters:
                print "meter uid: ", uid
                meter = models.Meter.objects.get(meter_uid=uid)
                account_meter = meter.accountmeter_set.get(status='active')
                account = account_meter.account
                address = account.address
                phase = address.address4
                print "phase: ", phase
                if ' ' in phase:
                    phase = phase.split(' ')[1]
                else:
                    phase = phase[0]
                block = address.address2
                lot = address.address3

                old_uid = meter.meter_uid
                new_uid = u'-'.join([old_uid,phase,block,lot])
                meter.meter_uid = new_uid
                meter.save()
                transaction.commit()



class BusinessDateLoader(object):

    def __init__(self, value):

        self.skip = False

        if value in ('now','today'):
            value = str(date.today())

        else:
            try:
                value = str(datetime.strptime(value,'%Y-%m-%d').date())
            except ValueError:
                print "ERROR: Invalid business_date VALUE (now, today) or FORMAT (YYYY-MM-DD)."

                self.skip = True

        self.value = value

    def load(self):

        if not self.skip:

            business_date, created = models.Config.objects.get_or_create(name='business_date', 
                    defaults= {'value':self.value, 
                    })

            if self.value != business_date.value:
                business_date.value = self.value

            business_date.last_updated = timezone.now()
            business_date.last_updated_by = 'BusinessDateLoader'
            business_date.save()


class ActivePeriodLoader(object):

    def __init__(self, value=None):

        self.skip = False

        print "value: ", value
        if value in ('now','today', None):
            value = str(date.today())

        else:
            try:
                value = str(datetime.strptime(value,'%Y-%m-%d').date())
            except ValueError:
                print "ERROR: Invalid business_date VALUE (now, today) or FORMAT (YYYY-MM-DD)."

                self.skip = True

        
        
        self.value = value

    def load(self):
        print "self.value: ", self.value
        if not self.skip:

            active_period = models.BillingSchedule.objects.get(start_date__lte = self.value, end_date__gte = self.value)
            config, created = models.Config.objects.get_or_create(name='active_period',
                defaults= {'value':'0',})
            config.value = str(active_period.pk)
            config.last_updated = timezone.now()
            config.last_updated_by = 'BusinessDateLoader'
            config.save()




class PaymentLoader(BaseLoader):
    Model = models.Payment

    def __init__(self, path, format='xlsx', headers=True):
        self.path = path 
        self.format = format
        self.headers = headers

    def load(self):

        from copy import copy
        load_details = {}
        processed = 0
        ignored = []
        error = []
        skipped = 0
        total = 0

        if self.format == 'csv':
            with open(self.path, 'r') as f:
                if self.format == 'csv': 
                    self.reader= csv.DictReader(f)
                else:
                    self.reader = []
                data = []
                for row in self.reader:
                    row['created_by'] = self.user
                    row['last_updated_by'] = self.user
                    data.append(self.Model(**row))
                print "loading", len(data), "for", self.Model
                self.Model.objects.bulk_create(data)
    
        elif self.format == 'xlsx':
            wb2 = load_workbook(self.path)
            print "sheets: ", wb2.get_sheet_names()
            ws = wb2.get_active_sheet()
            rows = ws.rows
            print "rows: ", rows
            load_details['rows'] = len(rows)

            fields = ['meter_uid', 'customer', 'address', 'amount', 'payment_date','payment_type','check_number']

            start = 0
            if self.headers:
                start = 1

            load_details['header'] = start
            
            for i in range(start, len(ws.rows)):
                print "cell %d:  %s", i, ws.cell(row=i, column=1)
                if (ws.cell(row=i, column=1).value is None) or (ws.cell(row=i, column=1).value.strip() == ''):
                    skipped+=1
                    continue                    
                else:
                    try:
                        row = {}
                        for field in fields:
                            print "field: ", field
                            print "cell: ", ws.cell(row=i,column=fields.index(field)).value
                            #row[i].decode('iso-8859-1').strip()
                            #print "decode: ", ws.cell(row=i,column=fields.index(field)).value.decode
                            row[field] = ws.cell(row=i,column=fields.index(field)).value

                        error.append(copy(row))
                        print "parsing.."
                        print "payment_date type: ", type(row['payment_date'])
                        print "payment_date: ", row['payment_date']
                        
                        if (row['payment_date'] is None):
                            print "Blank payment date.. ignoring row %d , row: %s " % (i, str(row))
                            ignored.append(row)
                            error.pop()
                            continue

                        meter_uid = row.pop('meter_uid')
                        print "meter_uid: ", meter_uid
                        #meter = models.Meter.objects.get(meter_uid=meter_uid)
                        address = row.pop('address').strip()
                        customer = row.pop('customer')
                        last_name, first_name = customer.split(",")
                        last_name = last_name.strip().upper()
                        first_name = first_name.strip().upper()
                        print "last_name: ", last_name
                        print "first_name: ", first_name
                        customer = models.Customer.objects.get(last_name=last_name, first_name=first_name)
                        address = models.Address.objects.get(address1=address)
                        account = models.Account.objects.get(customer=customer, address=address)
                        row['remarks'] = "Uploaded from file: "+ self.path.split("/")[-1]
                        check_number = row.pop('check_number')
                        if check_number is None:
                            row['check_number'] = ''
                        print "row, before addingl.. : ", row
                        account.add_payment(**row)
                        processed+=1
                        error.pop()
                        #d['account'] = account
                        
                        #data = Model(**d)
                        #data.save()
                    except Exception, e:
                        err = error.pop()
                        error.append("row: %s  error: % s" % (str(err), str(e)))
                        print "error: ",e

            
                        #print "Skipping row with Error", meter_uid, customer, address, e
            load_details['processed'] = processed
            load_details['error'] = len(error)
            load_details['skipped'] = skipped
            load_details['ignored'] = len(ignored)
            load_details['total'] = processed + len(ignored) + len(error)

            print "Load Results: " + str(load_details)
            if error:
                print "Errors!"
                print
                print

                for err in error:
                    print err

class AdjustmentLoader(BaseLoader):
    Model = models.Adjustment

    def __init__(self, path, format='xlsx', headers=True):
        self.path = path 
        self.format = format
        self.headers = headers

    def load(self):

        if self.format == 'csv':
            with open(self.path, 'r') as f:
                if self.format == 'csv': 
                    self.reader= csv.DictReader(f)
                else:
                    self.reader = []
                data = []
                for row in self.reader:
                    row['created_by'] = self.user
                    row['last_updated_by'] = self.user
                    data.append(self.Model(**row))
                print "loading", len(data), "for", self.Model
                self.Model.objects.bulk_create(data)
    
        elif self.format == 'xlsx':
            print "xlsx format! loading.."
            wb2 = load_workbook(self.path)
            print "sheets: ", wb2.get_sheet_names()
            ws = wb2.get_active_sheet()
            rows = ws.rows
            print "rows: ", rows
            

            fields = ['meter_uid', 'customer', 'address', 'amount', 'adjustment_date', 'adjustment_type']

            start = 0
            if self.headers:
                start = 1

            
            for i in range(start, len(ws.rows)):
                print "cell %d:  %s", i, ws.cell(row=i, column=1)
                if (ws.cell(row=i, column=1).value is None) or (ws.cell(row=i, column=1).value.strip() == ''):
                    continue
                else:
                    try:
                        row = {}
                        for field in fields:
                            print "field: ", field
                            print "cell: ", ws.cell(row=i,column=fields.index(field)).value
                            #row[i].decode('iso-8859-1').strip()
                            #print "decode: ", ws.cell(row=i,column=fields.index(field)).value.decode
                            row[field] = ws.cell(row=i,column=fields.index(field)).value

                        print "parsing.."
                        print "row: ", row
                        meter_uid = row.pop('meter_uid')
                        print "meter_uid: ", meter_uid
                        #meter = models.Meter.objects.get(meter_uid=meter_uid)
                        address = row.pop('address').strip()
                        customer = row.pop('customer')
                        last_name, first_name = customer.split(",")
                        last_name = last_name.strip().upper()
                        first_name = first_name.strip().upper()
                        print "last_name: ", last_name
                        print "first_name: ", first_name
                        customer = models.Customer.objects.get(last_name=last_name, first_name=first_name)
                        address = models.Address.objects.get(address1=address)
                        account = models.Account.objects.get(customer=customer, address=address)
                        row['description'] = "Initial balance uploaded from file: "+ self.path.split("/")[-1]
                        account.add_adjustment(**row)
                        #d['account'] = account
                        
                        #data = Model(**d)
                        #data.save()
                    except Exception, e:
                        print "error: ",e
                        #print "Skipping row with Error", meter_uid, customer, address, e





class Command(BaseCommand):
    args = '<poll_id poll_id ...>'
    help = 'Closes the specified poll for voting'
    option_list = BaseCommand.option_list + (
        make_option('--customer',            
            dest='customer',
            #default=False,
            help='load customer data into customer table'),
        make_option('--address',
            dest='address',
            #default=False,
            help='load  address into address table'),
        make_option('--rate',            
            dest='rate',
            #default=False,
            help='load rate into rate table'),
        make_option('--meter',            
            dest='meter',
            #default=False,
            help='load meter into meter table'),
        make_option('--ratecharge',            
            dest='ratecharge',
            #default=False,
            help='load ratecharge into ratecharge table'),
        make_option('--billingschedule',            
            dest='billingschedule',
            #default=False,
            help='load billingschedule into billingschedule table'),
        make_option('--accounttype',            
            dest='accounttype',
            #default=False,
            help='load accounttype into accounttype table'),
        make_option('--account',            
            dest='account',
            #default=False,
            help='load account into account table'),
        make_option('--meterread',            
            dest='meterread',
            #default=False,
            help='load meterread into meterread table'),
        make_option('--migrate',            
            dest='migrate',
            #default=False,
            help='migrate data'),

        make_option('--business-date',            
            dest='business_date',
            nargs=1,
            default=None,
            #default=False,
            help='update business-date'),

        make_option('--active-period',            
            dest='active_period',
            nargs=1,
            #default=False,
            help='determine active period'),

        make_option('--payment',            
            dest='payment',
            #default=False,
            help='load payments'),

        make_option('--adjustment',            
            dest='adjustment',
            #default=False,
            help='load adjustments'),

        make_option('--clear-financials',            
            dest='clear_financials',
            nargs=0,
            #default=False,
            help='clear financial tables'),

        )

    def handle(self, *args, **options):
        print "args: ", args
        print "options: ", options

        
        
        if options['clear_financials'] is not None:
            print "starting clearing.. "
            from django.db import connection
            from django.db import transaction
            from django.core.management import call_command

            cursor = connection.cursor()
            removed_tables = []
            exceptions = []

            Models = [models.Adjustment, models.PostedPayment, models.Payment, models.Bill,  models.Penalty, models.ReadCharge, models.FinancialTransaction,]

            # Delete Data
            for model in Models:
                try:
                    model.objects.all().delete()
                except Exception as e:
                    print "e: ", e

            # Drop the tables 
            for model in Models:
                try:
                    print "table: ", model._meta.db_table
                    model.objects.all().delete()
                    cursor.execute('DROP TABLE %s' % model._meta.db_table)
                    print "Dropped table %s from model %s" % (model._meta.db_table, model.__name__)
                except Exception as e:
                    print "e: ", e
                    exceptions.append([model._meta.db_table, str(e)])
                removed_tables.append(model._meta.db_table)
            
            cursor.close()
            
            transaction.commit()
            
            print "starting sync.."

            call_command('syncdb')


            print "done with sync.."

        if options['customer']:
            print 'loading customer'
            loader = CustomerLoader(options['customer'])
            loader.load()

        if options['address']:
            print 'loading address'
            loader = AddressLoader(options['address'])
            loader.load()

        if options['rate']:
            print 'loading rate'
            loader = RateLoader(options['rate'])
            loader.load()

        if options['meter']:
            print 'loading meter'
            loader = MeterLoader(options['meter'])
            loader.load()

        if options['ratecharge']:
            print 'loading ratecharge'
            loader = RateChargeLoader(options['ratecharge'])
            loader.load()

        if options['billingschedule']:
            print 'loading billingschedule'
            loader = BillingScheduleLoader(options['billingschedule'])
            loader.load()

        if options['accounttype']:
            print 'loading accounttype'
            loader = AccountTypeLoader(options['accounttype'])
            loader.load()

        if options['account']:
            print 'loading account'
            loader = AccountLoader(options['account'])
            loader.load()

        if options['meterread']:
            print 'loading meterread'
            loader = MeterReadLoader(options['meterread'])
            loader.load()   

        if options['migrate']:
            print 'loading migrate'
            loader = FelizanaLoader(options['migrate'])
            loader.load()

        if options['business_date']:
            print "updating business-date"
            loader = BusinessDateLoader(options['business_date'])
            loader.load()

        if options['payment']:
            print 'loading payment'
            loader = PaymentLoader(options['payment'])
            loader.load()

        if options['adjustment']:
            print 'loading adjustment'
            loader = AdjustmentLoader(options['adjustment'])
            loader.load()

        if options['active_period'] == None:
            pass
        else:
            print "active_period: ", options['active_period']
            print "determining active period"
            loader = ActivePeriodLoader(options['active_period'])
            loader.load()




        





    # if argv is None:
    #   argv = sys.argv

    # parser = argparse.ArgumentParser(description='Load billing data')
    # parser.add_argument('--customer', dest='customer', type=argparse.FileType('r'),
    #                 help='load customer data into customer table')

    # args = parser.parse_args(argv.split())
    # if args.customer:
    #   print args.customer
