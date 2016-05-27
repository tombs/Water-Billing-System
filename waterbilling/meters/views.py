# -*- coding: utf-8 -*-
# Create your views here.
import json
import csv
from decimal import Decimal
from django.utils import timezone
from django.views.generic import ListView, DetailView, TemplateView
from django.views.generic.edit import CreateView
from django.http import HttpResponseRedirect, HttpResponse
from django.db.models import Sum, Max, F
from eztables.views import DatatablesView
from .models import Meter, MeterRead, Config, BillingSchedule, Customer, Account, Address, Task
from .forms import MeterReadUploadForm, MeterReadForm
from core.views import login_required, ajax_view
from core.utils import UnicodeWriter
from datetime import datetime, time
from openpyxl import load_workbook, Workbook
from tempfile import mkstemp
import os
from openpyxl.writer.excel import save_virtual_workbook

@login_required
class MeterReadListView(TemplateView):
    def get_context_data(self, **kwargs):
        context = super(MeterReadListView, self).get_context_data(**kwargs)
        period = Config.objects.get(name='active_period').value
        bs = BillingSchedule.objects.get(pk=period)
        business_date = Config.objects.get(name='business_date').value
        business_date = datetime.strptime(business_date,'%Y-%m-%d')
        business_date = business_date.strftime('%b %d, %Y')
        context['period'] = str(bs)
        context['usage'] = bs.reading_start_date.strftime("%b %d, %Y") + " - " + bs.reading_end_date.strftime("%b %d, %Y")
        context['business_date'] = business_date
        return context

@login_required
@ajax_view
class MeterReadDatatablesView(DatatablesView):
    model = MeterRead
    fields = (
        'id',
        'read_date',         
        '{account__customer__last_name}, {account__customer__first_name}',
        'account__address__address4',
        'meter__status',
        'previous_reading',
        'current_reading',
        'usage',
        'account__id',        
        )

@login_required
class MeterReadDetailView(DetailView):

    model = MeterRead
    #queryset = Account.objects.filter(account_type__account_type='Residential')
    context_object_name = 'meter_read_detail'

    def get_context_data(self, **kwargs):
        context = super(MeterReadDetailView, self).get_context_data(**kwargs)
        print context
        #context['bill_detail'] = context['account_detail'].bill_set.latest('bill_date')
        context['account'] = context['meter_read_detail'].meter.account_set.latest('last_updated')
        #context['now'] = timezone.now()
        return context

@login_required
@ajax_view
class MeterReadUploadView(CreateView):
    form_class = MeterReadUploadForm
    model = MeterRead

    def render_to_json_response(self, context, **response_kwargs):
        data = json.dumps(context)
        response_kwargs['content_type'] = 'application/json'
        return HttpResponse(data, **response_kwargs)

    # def get(self, request, *args, **kwargs):
    #   print "This POST"
    #   form = self.form_class(initial=self.initial)
    #   return render(request, self.template_name, {'form': form})

    def post(self, request, headers=True, *args, **kwargs):
        print " ---- Meter Reads Upload Starting! "
        form = self.form_class(request.POST, request.FILES)        
        print "REQUEST", request.POST, request.FILES, request.is_ajax()
        print "FORM", form.is_valid(), form.errors
        print "--- SECONDS: ", request.POST.get('pk')

        today = datetime.strptime(Config.objects.get(name='business_date').value, '%Y-%m-%d').date()
        period = Config.objects.get(name='active_period').value
        bs = BillingSchedule.objects.get(pk=period)

        self.headers = headers
        load_details = {}
        if form.is_valid():
            f = request.FILES.get('file')
            wb2 = load_workbook(filename=f)
            print "loaded!"
            print "sheets: ", wb2.get_sheet_names()
            ws = wb2.get_active_sheet()
            rows = ws.rows
            print "rows: ", rows
            load_details['rows'] = len(rows)

            fields = ['meter_uid', 'customer', 'address', 'read_date', 'read_time','previous_reading','current_reading']

            start = 0
            if self.headers:
                start = 1

            load_details['header'] = start

            from copy import copy
            load_details = {}
            processed = 0
            ignored = []
            error = []
            skipped = 0
            total = 0
            rows = 0

            data = []
            errors = []


            period = Config.objects.get(name='active_period').value
            bs = BillingSchedule.objects.get(pk=period)

            task_id = request.POST.get('pk')
            
            print "task_id: ", task_id
            print "----- creating a task.."
            task, task_created = Task.objects.get_or_create(name='Upload Meter Reads',type='meter', business_date=today, task_id=task_id)
            print "----- task created! : ", task_created

            counter = start
            if task_created:
                task.task_id = task_id
                task.created_by = request.user.username
                accounts =  self.model.objects.exclude(status='inactive') 
                total = len(ws.rows)
                task.jobs_total = total
                task.jobs_done = counter
                task.status = 'in progress'
                task.save()



            for i in range(start, len(ws.rows)):
                print "cell %d:  %s", i, ws.cell(row=i, column=1)                
                if (ws.cell(row=i, column=1).value is None) or (ws.cell(row=i, column=1).value.strip() == ''):
                    skipped+=1
                    counter+=1                    
                    task.jobs_done = counter
                    task.save()

                    continue                    
                else:
                    try:
                        row = {}
                        for field in fields:
                            print "field: ", field
                            print "cell: ", ws.cell(row=i,column=fields.index(field)).value
                            #row[i].decode('iso-8859-1').strip()
                            #print "decode: ", ws.cell(row=i,column=fields.index(field)).value.decode
                            row[field] = ws.cell(row=i,column=fields.index(field)).value

                        error.append(copy(row))
                        print "parsing.."
                        
                        if (row['current_reading'] is None):
                            print "Blank current reading.. ignoring row %d , row: %s " % (i, str(row))
                            ignored.append(row)
                            error.pop()
                            continue

                        meter_uid = row.pop('meter_uid')
                        print "meter_uid: ", meter_uid
                        #meter = models.Meter.objects.get(meter_uid=meter_uid)
                        address = row.pop('address').strip()
                        if row.get("read_date"):
                            read_date = row.pop('read_date')
                            read_date = read_date.strftime('%Y-%m-%d')
                        else:
                            read_date = None
                        print "read_date: ", read_date
                        customer = row.pop('customer')
                        last_name, first_name = customer.split(",")
                        last_name = last_name.strip().upper()
                        first_name = first_name.strip().upper()
                        print "last_name: ", last_name
                        print "first_name: ", first_name
                        customer = Customer.objects.get(last_name=last_name, first_name=first_name)
                        address = Address.objects.get(address1=address)
                        account = Account.objects.get(customer=customer, address=address)
                        print "row, before adding.. : ", row
                        
                        account.add_meter_read(current_reading=row['current_reading'], read_date=read_date, period=bs.id)

                        processed+=1
                        error.pop()
                        #d['account'] = account
                        
                        #data = Model(**d)
                        #data.save()
                    except Exception, e:
                        err = error.pop()
                        error.append("row: %s  error: %s" % (str(err), str(e)))
                        print "error: ",e

                    counter+=1                    
                    task.jobs_done = counter
                    task.save()

            
                        #print "Skipping row with Error", meter_uid, customer, address, e
            load_details['processed'] = processed
            load_details['error'] = len(error)
            load_details['skipped'] = skipped
            load_details['ignored'] = len(ignored)
            load_details['total'] = processed + len(ignored) + len(error)

            print "Load Results: " + str(load_details)

            if error:
                print "Errors!"
                print
                print

                for err in error:
                    print err


            f.close()

            print "-- finished uploading meter reads.. "
            task.status = 'completed'     
            task.result = str(load_details)           
            task.save()
            print "-- status has been saved! "




            if self.request.is_ajax():
                data = {
                    'data': {'count': len(data)},
                    'errors': errors
                }
                print "response data", data
                if errors:
                    return self.render_to_json_response(data, status=409) #conflict
                else:
                    data['message'] = 'succesful'
                    return self.render_to_json_response(data)

            return HttpResponseRedirect('/meters/')

        else:
            if self.request.is_ajax():
                result =  self.render_to_json_response(form.errors, status=400)
                print 'RESULT ', result
                return result
            return HttpResponseRedirect('/meters/')

@login_required
class MeterReadDownloadFileView(DetailView):

    queryset = MeterRead.objects.filter(id__in=Meter.objects.annotate(
                    last_meterread_id=Max('meterread')).values_list('last_meterread_id')).values(
            'meter__meter_uid',
            'account__address__address1',
            'account__customer__last_name',
            'account__customer__first_name',
            #'current_reading').order_by('account__address__address4','account__address__address2', 'account__address__address3')
            'current_reading').order_by('account__created')
                

    def get(self,request, *args, **kwargs):
        #https://docs.djangoproject.com/en/dev/howto/outputting-csv/
        period = Config.objects.get(name='active_period').value
        bs = BillingSchedule.objects.get(pk=period)

        data = self.queryset.all()
        header = ['meter_uid','customer','address','read_date','read_time','previous_reading','current_reading']
        kwargs['content_type'] = 'text/csv'

        filename = "MeterReadings_" + bs.next.reading_start_date.strftime("%b_%d_%Y") + "-" + bs.next.reading_end_date.strftime("%b_%d_%Y")  + ".xlsx"      
        #  -- CONVERSION START
        wb = Workbook()
        ws = wb.active
        cur = 0 #current row
        #set headers
        for field in header:
            cell = ws.cell(row=cur,column=header.index(field))
            cell.value = field

        cur+=1

        #set data
        for row in data:
            cell = ws.cell(row=cur,column=0)
            cell.value = row['meter__meter_uid']
            cell = ws.cell(row=cur,column=1)
            cell.value = row['account__customer__last_name'] + u',' + row['account__customer__first_name']
            cell = ws.cell(row=cur,column=2)
            cell.value = row['account__address__address1']
            cell = ws.cell(row=cur,column=3)
            cell.value = str(bs.reading_end_date)
            cell = ws.cell(row=cur,column=4)
            cell.value = ''
            cell = ws.cell(row=cur,column=5)
            cell.value = unicode(row['current_reading'])
            cell = ws.cell(row=cur,column=6)
            cell.value = ''

            cur+=1


        response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="' + filename +'"'

        return response



@login_required
@ajax_view
class MeterReadUpdateView(DetailView):

    model = MeterRead
     
    def render_to_json_response(self, context, **response_kwargs):
        data = json.dumps(context)
        response_kwargs['content_type'] = 'application/json'
        return HttpResponse(data, **response_kwargs)

    def post(self, request, *args, **kwargs):
        print "This POST", request.POST
        request = request.POST
        meterread = self.model.objects.get(pk = request.get('id'))
        
        try:
            if request.get('current_reading'):
                meterread.current_reading = request.get('current_reading')
            if request.get('previous_reading'):
                meterread.previous_reading = request.get('previous_reading')
            meterread.usage = Decimal(meterread.current_reading) - Decimal(meterread.previous_reading)
            meterread.update()
            data = {'status': 200, 'data': meterread.status, 'msg': 'Updated Meter Read with ID: ' + str(meterread.pk)}
            status = 200
        except Exception, e:
            data = {'status': 200, 'data': meterread.status, 'msg': 'Unable to update Meter Read with ID: ' + request.get('pk') }
            status = 400

        return self.render_to_json_response(data, status=status)


@login_required
@ajax_view
class MeterAddView(DetailView):

    model = Meter
     
    def render_to_json_response(self, context, **response_kwargs):
        data = json.dumps(context)
        response_kwargs['content_type'] = 'application/json'
        return HttpResponse(data, **response_kwargs)

    def post(self, request, *args, **kwargs):
        print "This POST", request.POST
        request = request.POST
        print "request: ", str(request)

        period = Config.objects.get(name='active_period').value
        bs = BillingSchedule.objects.get(pk=period)
        business_date = Config.objects.get(name='business_date').value
        business_date = datetime.strptime(business_date,'%Y-%m-%d')

        account = Account.objects.get(pk=request.get('account'))
        current_reading = request.get('current_reading')
        new_meter =request.get('new_meter')
        new_meter_reading = Decimal(request.get('new_meter_reading'))

        meter = Meter.objects.filter(meter_uid=new_meter)
        print "the meter: ", meter
        if meter.count() == 0:
            meter = Meter.objects.create(meter_uid = new_meter, 
                                date_installed = business_date, status='active') 
        else:
            meter = meter.get()

        
        try:
            if account.has_bill():
                print "Bill exists.. moving to next period"
                bs = bs.next
                account.add_meter_read(current_reading=current_reading,period=bs.id)
            else:
                print "No bills yet.. Adding meter read"
                account.add_meter_read(current_reading=current_reading,period=bs.id)
            
            account.add_account_meter(meter.pk) 
            print "Done with Add Account Meter"

            # Create a dummy meter read for the new meter
            meterread, created = MeterRead.objects.get_or_create(meter = meter, account = account, billing_schedule = bs,
                                    read_date=business_date,
                                    defaults={
                                                'read_time':time(),
                                                'current_reading': new_meter_reading,
                                                'previous_reading':new_meter_reading,
                                                'usage':Decimal("0.0")
                                            })
            print "meterread", meterread, created

            data = {'status': 200, 'data': meter.status, 'msg': 'Added new meter with ID: ' + str(meter.meter_uid)}
            status = 200
        except Exception, e:
            print "e: ", str(e)
            data = {'status': 200, 'data': meter.status, 'msg': 'Unable to add meter with : ' + request.get('new_meter') }
            status = 400

        return self.render_to_json_response(data, status=status)
